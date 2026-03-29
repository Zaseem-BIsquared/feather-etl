"""Create Excel test fixtures matching CSV fixture data.

Usage: uv run python scripts/create_excel_fixture.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl


FIXTURE_DIR = Path(__file__).parent.parent / "tests" / "fixtures" / "excel_data"


def main() -> None:
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)

    # orders.xlsx — 5 rows
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["order_id", "customer_id", "product", "quantity", "price"])
    ws.append([1, 101, "Widget A", 2, 10.0])
    ws.append([2, 102, "Widget B", 1, 20.0])
    ws.append([3, 101, "Widget A", 3, 10.0])
    ws.append([4, 103, "Widget C", 1, 30.0])
    ws.append([5, 102, "Widget B", 2, 20.0])
    wb.save(FIXTURE_DIR / "orders.xlsx")

    # customers.xlsx — 4 rows
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id", "name", "city", "active"])
    ws.append([101, "Acme Corp", "Mumbai", True])
    ws.append([102, "Beta Ltd", "Delhi", True])
    ws.append([103, "Gamma Inc", "Chennai", False])
    ws.append([104, "Delta Co", "Bangalore", True])
    wb.save(FIXTURE_DIR / "customers.xlsx")

    # products.xlsx — 3 rows
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id", "name", "category", "price"])
    ws.append([1, "Widget A", "Standard", 10.0])
    ws.append([2, "Widget B", "Standard", 20.0])
    ws.append([3, "Widget C", "Premium", 30.0])
    wb.save(FIXTURE_DIR / "products.xlsx")

    print(f"Created Excel fixtures in {FIXTURE_DIR}")
    print("  orders.xlsx: 5 rows")
    print("  customers.xlsx: 4 rows")
    print("  products.xlsx: 3 rows")


if __name__ == "__main__":
    main()
